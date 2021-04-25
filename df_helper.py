import pandas as pd
import streamlit as st
from openpyxl import load_workbook

def get_df(file):
    # get extension and read file
    extension = file.name.split('.')[1].upper()
    
    ### PICKLE ###
    if extension == 'PICKLE':
        df = pd.read_pickle(file)
    else:
        # Extension parameters expander
        expander_ext = st.sidebar.beta_expander(f'{extension} parameters')
        st.sidebar.text(" \n") #break line
    ### CSV ###
    if extension == 'CSV':
        # encoding
        encodings = pd.read_csv('python_encodings.csv')['Encoding'].values
        encoding = expander_ext.selectbox('Select Encoding', encodings, 95)
        # bad lines
        bad_lines_dict = {'Drop':False, 'Break':True}
        i = expander_ext.selectbox('Bad lines', [*bad_lines_dict])
        # data frame
        df = pd.read_csv(file, encoding = encoding, error_bad_lines=bad_lines_dict[i])
    
    ### EXCEL ###
    elif extension == 'XLSX':
        wb = load_workbook(file)
        # Work sheet
        selected_ws = expander_ext.selectbox('Select Sheet', wb.sheetnames)
        # Start from row
        max_row = wb.worksheets[wb.sheetnames.index(selected_ws)].max_row
        start_r = expander_ext.slider('Start from row', 1, max_row)
        # data frame
        df = pd.read_excel(file, engine='openpyxl', sheet_name=selected_ws, header=start_r-1)

    return df


def convert_dtypes(df, types, new_types):
    for i, col in enumerate(df.columns):
        # get current column new data type
        new_type = types[new_types[i]]
        if new_type: # check if none
            try: # try to convert
                df[col] = df[col].astype(new_type)
            except: # error msg
                st.write('Could not convert', col, 'to', new_types[i] )
    return df

def handle_nulls(df, null_dict, n_dict):
    for i, col in enumerate(df.columns):
        # check if column is in the null dictionary
        if i in n_dict:
            # get null procedure from dictionary
            # 0 - Remove na
            # 1 - Text
            # 2 - Mean
            # 3 - Median 
            null_proc = null_dict[n_dict[i]]
            try:
                if null_proc == 0:
                    df = df[df[col].notna()]
                elif null_proc == 1:
                    df[col] = df[col].fillna('Not Available')
                elif null_proc == 2:
                    df[col] = df[col].fillna(df[col].mean())
                elif null_proc == 3:
                    df[col] = df[col].fillna(df[col].median())
                elif null_proc == 4:
                    df[col] = df[col].fillna(0)
            except:
                st.write('Could not convert', col, 'nulls to', n_dict[i])
    return df

def handle_duplicates(df, duplicates_dict, action):
    try:
        df = df[~df.duplicated(keep=duplicates_dict[action])]
    except:
        pass
    return df